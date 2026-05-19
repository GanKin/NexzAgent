"""
市场概览 Excel 解析器

使用 openpyxl 解析体系 B（中文列名）的 Excel 数据文件。
"""

import re
from typing import Any, Dict, List, Optional, Tuple

from app.utils.logger import get_logger

logger = get_logger(__name__)

# 体系 B 列名到数据库字段的映射
COLUMN_MAPPING = {
    '代码': 'symbol',
    '标的名称': 'name',
    '相对强度': 'relative_strength',
    '强度动量': 'strength_momentum',
    '早期转折': 'early_reversal',
    '当前比价状态': 'relative_price_status',
    '当前比价状态持续时间': 'relative_price_duration',
    '当前比价状态涨幅': 'relative_price_return',
    '此前比价状态': 'prev_relative_price_status',
    '此前比价状态持续时间': 'prev_relative_price_duration',
    '日级别趋势': 'd_trend',
    '日级别趋势持续时间': 'd_trend_duration',
    '周级别趋势': 'w_trend',
    '周级别趋势持续时间': 'w_trend_duration',
    '月级别趋势': 'm_trend',
    '月级别趋势持续时间': 'm_trend_duration',
    '收盘价对比60日位置': 'price_position_60d',
    '当前杠杆资金状态': 'leverage_status',
    '当前杠杆资金状态持续时间': 'leverage_duration',
    '当前杠杆资金状态涨幅': 'leverage_return',
    '此前杠杆资金状态': 'prev_leverage_status',
    '杠杆资金数值': 'leverage_value',
    '杠杆资金相比前日变动': 'leverage_change',
}

# 数值类型字段（需要转 float/int）
_NUMERIC_FIELDS = {
    'relative_strength', 'strength_momentum', 'early_reversal',
    'relative_price_duration', 'relative_price_return',
    'prev_relative_price_duration',
    'd_trend_duration', 'w_trend_duration', 'm_trend_duration',
    'price_position_60d',
    'leverage_duration', 'leverage_return',
    'leverage_value', 'leverage_change',
}

# 趋势字段（空值兜底为 '无趋势'）
_TREND_FIELDS = {'d_trend', 'w_trend', 'm_trend'}

# 状态字段（空值兜底为 ''）
_STATUS_FIELDS = {
    'relative_price_status', 'prev_relative_price_status',
    'leverage_status', 'prev_leverage_status',
}


class MarketOverviewParser:
    """市场概览 Excel 解析器 — 体系 B（中文列名）"""

    def parse(self, file_path: str) -> Tuple[List[Dict], List[Dict]]:
        """
        解析 Excel 文件，返回 (records, errors)。

        使用 openpyxl 逐行读取，避免 pandas 的 NaN 类型问题。

        Args:
            file_path: Excel 文件路径

        Returns:
            records: 解析后的数据记录列表（每条为字段名→值的字典）
            errors: 解析错误列表（每条包含 row, field, error）
        """
        records = []
        errors = []

        try:
            from openpyxl import load_workbook
        except ImportError:
            logger.error("openpyxl 未安装，无法解析 Excel 文件")
            return [], [{'row': 0, 'field': 'dependency', 'error': 'openpyxl not installed'}]

        try:
            wb = load_workbook(file_path, data_only=True, read_only=True)
            ws = wb.active

            # 读取第一行作为 header
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header_row:
                return [], [{'row': 0, 'field': 'header', 'error': 'Excel 文件为空'}]

            headers = [str(h).strip() if h else '' for h in header_row]

            if not self.validate_header(headers):
                return [], [{'row': 0, 'field': 'header', 'error': f'非体系 B 格式，缺少必要列。实际列: {headers[:5]}...'}]

            # 构建 header → db_field 映射
            col_map = {}
            for i, h in enumerate(headers):
                if h in COLUMN_MAPPING:
                    col_map[i] = COLUMN_MAPPING[h]

            # 逐行读取数据
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                record = {}
                row_has_symbol = False

                for col_idx, db_field in col_map.items():
                    if col_idx >= len(row):
                        continue

                    raw_value = row[col_idx]
                    value = self._normalize_value(db_field, raw_value)
                    record[db_field] = value

                    if db_field == 'symbol' and value:
                        row_has_symbol = True

                # 跳过没有 symbol 的行
                if not row_has_symbol:
                    continue

                # 确保所有数值字段有默认值
                for field in _NUMERIC_FIELDS:
                    if field not in record or record[field] is None:
                        record[field] = 0
                for field in _TREND_FIELDS:
                    if field not in record or not record[field]:
                        record[field] = '无趋势'
                for field in _STATUS_FIELDS:
                    if field not in record or not record[field]:
                        record[field] = ''

                records.append(record)

            wb.close()

        except Exception as e:
            logger.error(f"Excel 解析失败: {e}", exc_info=True)
            errors.append({'row': 0, 'field': 'parse', 'error': str(e)[:200]})

        return records, errors

    @staticmethod
    def validate_header(headers: List[str]) -> bool:
        """
        验证 header 是否属于体系 B。

        检查是否包含 '代码' 和 '日级别趋势' 关键列。
        """
        header_set = set(headers)
        return '代码' in header_set and '日级别趋势' in header_set

    @staticmethod
    def _normalize_value(db_field: str, raw_value: Any) -> Any:
        """
        根据 db_field 类型规范化值。

        - 趋势字段空值 → '无趋势'
        - 数值字段 → float/int，空值 → 0
        - 文本字段 → str
        """
        if raw_value is None:
            if db_field in _TREND_FIELDS:
                return '无趋势'
            if db_field in _NUMERIC_FIELDS:
                return 0
            if db_field in _STATUS_FIELDS:
                return ''
            return ''

        # 趋势字段：直接转字符串
        if db_field in _TREND_FIELDS:
            val = str(raw_value).strip()
            return val if val and val != 'None' else '无趋势'

        # 状态字段：直接转字符串
        if db_field in _STATUS_FIELDS:
            val = str(raw_value).strip()
            return val if val and val != 'None' else ''

        # symbol 和 name：字符串
        if db_field in ('symbol', 'name'):
            val = str(raw_value).strip()
            return val if val and val != 'None' else ''

        # 数值字段
        if db_field in _NUMERIC_FIELDS:
            try:
                if isinstance(raw_value, (int, float)):
                    return raw_value
                val = str(raw_value).strip().replace('%', '').replace(',', '')
                if not val or val == 'None' or val == '-':
                    return 0
                return float(val)
            except (ValueError, TypeError):
                return 0

        # 其他字段：字符串
        return str(raw_value).strip()
